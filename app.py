import os
import pandas as pd
import difflib
import re
from supabase import create_client, Client
from werkzeug.utils import secure_filename
from datetime import datetime
from flask import Flask, request, jsonify, render_template_string, redirect
from flask import redirect, url_for
from openpyxl import load_workbook
from collections import defaultdict
from flask import session, redirect, url_for
from werkzeug.security import check_password_hash, generate_password_hash
from functools import wraps
from flask import render_template, request

# === Supabase Config ===
SUPABASE_URL = "https://wmthdsalqsrdiwxbmfey.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6IndtdGhkc2FscXNyZGl3eGJtZmV5Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NDQyNjE2NTEsImV4cCI6MjA1OTgzNzY1MX0.3_3iBmNy9VM5BRycEKLvTfBBCRQkjJF7kP0EN-VZH68"
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
# Simple user store (you can also use .env or Supabase if needed)
users = {
    "ailianyvette@gmail.com": generate_password_hash("jojo16022001"),
    "kuoyen23@yahoo.com": generate_password_hash("jojo16022001")
}

# === Flask App ===
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "supersecretkey123")
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)



def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get("user"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/login', methods=['GET', 'POST'])
def login():
    message = None
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')

        if email in users and check_password_hash(users[email], password):
            session['user'] = email
            return redirect(url_for('search_form'))
        else:
            message = "❌ Invalid credentials."

    return render_template_string("""
    <html>
        <head>
            <title>Login</title>
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
        </head>
        <body class="container py-5">
            <h2 class="mb-4">🔐 Login to CPSApp</h2>
            {% if message %}
                <div class="alert alert-danger">{{ message }}</div>
            {% endif %}
            <form method="post">
                <div class="mb-3">
                    <label>Email</label>
                    <input type="email" name="email" class="form-control" required>
                </div>
                <div class="mb-3">
                    <label>Password</label>
                    <input type="password" name="password" class="form-control" required>
                </div>
                <button class="btn btn-primary" type="submit">Login</button>
            </form>
        </body>
        <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>

    </html>
    """, message=message)

@app.route('/logout')
def logout():
    session.pop('user', None)
    return redirect(url_for('login'))


# === Upload Endpoint ===
@app.route('/upload', methods=['POST'])
@login_required
def upload_product_api():  #
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    try:
        df = pd.read_excel(filepath, sheet_name='Master')
        wb = load_workbook(filepath, data_only=True)

        for _, row in df.iterrows():
            code_raw = str(row['Code'])
            code = code_raw.replace("Code:", "").strip()

            product_tab = row['Description'].strip()
            if product_tab in wb.sheetnames:
                sheet = wb[product_tab]
                print(f"[DEBUG] Tab: {product_tab}")
                for row in [5, 6, 7]:
                    for col in ['F', 'G', 'H', 'I', 'J']:
                        cell_ref = f'{col}{row}'
                        cell = sheet[cell_ref]
                        print(f"[DEBUG] {cell_ref}: {cell.value if cell else 'None'}")
                measurements = extract_measurements(sheet)

            else:
                measurements = ""

            data = {
                "design_number": int(row['Design Number']),
                "description": row['Description'],
                "price": float(row['Price']),
                "total_quantity": int(row['Total Quantity']),
                "color": row['Color'],
                "code": code,
                "upload_date": datetime.utcnow().isoformat(),
                "source_file": filename,
                "measurements": measurements
            }

            supabase.table("products").insert(data).execute()

        return jsonify({"message": "Upload and insert successful"})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# === Run the App ===
from flask import render_template_string

@app.route('/search', methods=['GET'])
def search():
    code = request.args.get('code')
    description = request.args.get('description')

    query = supabase.table("products").select("*")

    if code:
        query = query.ilike("code", f"%{code}%")
    elif description:
        query = query.ilike("description", f"%{description}%")
    else:
        return "Please provide a code or description", 400

    result = query.execute()
    data = result.data

    # HTML table template
    html_template = """
    <html>
        <head><title>Search Results</title></head>
        <body>
            <h2>Search Results</h2>
            {% if data %}
            <table border="1" cellpadding="8">
                <tr>
                    {% for key in data[0].keys() %}
                        <th>{{ key }}</th>
                    {% endfor %}
                </tr>
                {% for row in data %}
                    <tr>
                    {% for value in row.values() %}
                        <td>{{ value }}</td>
                    {% endfor %}
                    </tr>
                {% endfor %}
            </table>
            {% else %}
                <p>No matching results found.</p>
            {% endif %}
        </body>
    </html>
    """
    return render_template_string(html_template, data=data)

def extract_measurements(sheet):
    values = []
    for row in [5, 6, 7]:
        cell = sheet[f'H{row}']
        if cell and cell.value:
            text = str(cell.value).strip()

            replacements = {
                r'\bptp\b': 'PTP',
                r'\bhip\b': 'Hip',
                r'\bhips\b': 'Hip',
                r'\bwaist\b': 'Waist',
                r'\blength\b': 'Length',
                r'\bl\b': 'Length',
                r'\bw\b': 'Waist',
                r'\bh\b': 'Hip',
                r'\binner\b': 'Inner',
                r'\bouter\b': 'Outer'
            }

            for pattern, replacement in replacements.items():
                text = re.sub(pattern, replacement, text, flags=re.IGNORECASE)
            text = re.sub(r'\s*,\s*', ', ', text)
            text = re.sub(r'\s+', ' ', text)

            values.append(text)

    return ', '.join(values)



@app.route('/upload-form', methods=['GET', 'POST'])
@login_required
def upload_form():
    message = None

    if request.method == 'POST':
        print("[DEBUG] Upload form POST triggered",flush=True)
        if 'file' not in request.files or request.files['file'].filename == '':
            message = "No file selected."
        else:
            print("[DEBUG] File selected: proceeding to save and process...",flush=True)
            file = request.files['file']
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)

            try:
                print("[DEBUG] Starting full upload processing...",flush=True)

                # Pack_List tab
                try:
                    print("[DEBUG] Processing Pack_List tab...",flush=True)
                    packlist_df = pd.read_excel(filepath, sheet_name='Pack_List')
                    print("[DEBUG] ✅ Loaded Pack_List",flush=True)
                    for i, row in packlist_df.iterrows():
                        row_dict = row.dropna().to_dict()
                        supabase.table("packlist").insert({
                            "source_file": filename,
                            "row_index": i,
                            "row_data": row_dict
                        }).execute()
                except Exception as e:
                    print(f"[DEBUG] ❌ Failed to load or save Pack_List: {str(e)}",flush=True)

                # Bill tab
                try:
                    print("[DEBUG] Attempting to read Bill tab...",flush=True)
                    bill_df = pd.read_excel(filepath, sheet_name='Bill')
                    print("[DEBUG] ✅ Loaded Bill",flush=True)
                    for i, row in bill_df.iterrows():
                        row_dict = row.dropna().to_dict()
                        print(f"[DEBUG] Bill row {i}: {row_dict}",flush=True)
                        supabase.table("bills").insert({
                            "source_file": filename,
                            "row_index": i,
                            "row_data": row_dict
                        }).execute()
                except Exception as e:
                    print(f"[DEBUG] ❌ Failed to load or save Bill tab: {str(e)}",flush=True)

                # Master tab
                df = pd.read_excel(filepath, sheet_name='Master')
                wb = load_workbook(filepath, data_only=True)

                for _, row in df.iterrows():
                    code_raw = str(row['Code'])
                    code = code_raw.replace("Code:", "").strip()
                    product_tab = row['Description'].strip()

                    tab_match = difflib.get_close_matches(product_tab, wb.sheetnames, n=1, cutoff=0.6)
                    if tab_match:
                        sheet = wb[tab_match[0]]
                        measurements = extract_measurements(sheet)
                    else:
                        measurements = ""

                    data = {
                        "design_number": int(row['Design Number']),
                        "description": row['Description'],
                        "price": float(row['Price']),
                        "total_quantity": int(row['Total Quantity']),
                        "color": row['Color'],
                        "code": code,
                        "upload_date": datetime.utcnow().isoformat(),
                        "source_file": filename,
                        "measurements": measurements
                    }

                    supabase.table("products").insert(data).execute()

                message = "✅ Upload and insert successful!"

            except Exception as e:
                print(f"[DEBUG] ❌ Outer upload error: {str(e)}",flush=True)
                message = f"❌ Upload failed: {str(e)}"

    return render_template_string("""
    <html>
        <head>
            <title>Upload Product Excel</title>
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
            <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
        </head>
        <body class="container py-5">
           <nav class="navbar navbar-expand-lg navbar-dark bg-dark mb-4">
                <div class="container-fluid">
                    <a class="navbar-brand" href="/search-form">🧾 CPSApp</a>
                    <button class="navbar-toggler" type="button" data-bs-toggle="collapse" data-bs-target="#mainNavbar" aria-controls="mainNavbar" aria-expanded="false" aria-label="Toggle navigation">
                    <span class="navbar-toggler-icon"></span>
                    </button>
                    
                    <div class="collapse navbar-collapse" id="mainNavbar">
                    <ul class="navbar-nav me-auto mb-2 mb-lg-0">
                        <li class="nav-item"><a class="nav-link {% if request.path == '/upload-form' %}active{% endif %}" href="/upload-form">📤 Upload</a></li>
                        <li class="nav-item"><a class="nav-link {% if request.path == '/search-form' %}active{% endif %}" href="/search-form">🔍 Search & Delete</a></li>
                        <li class="nav-item"><a class="nav-link {% if request.path == '/view-packlist' %}active{% endif %}" href="/view-packlist">📦 Pack_List</a></li>
                        <li class="nav-item"><a class="nav-link {% if request.path == '/invoice' %}active{% endif %}" href="/invoice">🧾 Invoice</a></li>
                        <li class="nav-item"><a class="nav-link {% if request.path == '/generate-invoice' %}active{% endif %}" href="/generate-invoice">🧾 Generate Invoice</a></li>
                    </ul>
                    {% if session.get("user") %}
                    <div class="d-flex align-items-center">
                        <span class="navbar-text text-white me-3">👋 {{ session['user'] }}</span>
                        <a href="/logout" class="btn btn-outline-light btn-sm">Logout</a>
                    </div>
                    {% endif %}
                    </div>
                </div>
            </nav>


            <h2 class="mb-4">📤 Upload Product Excel File</h2>
            {% if message %}
                <div class="alert alert-info">{{ message }}</div>
            {% endif %}
            <form method="post" enctype="multipart/form-data">
                <div class="mb-3">
                    <input class="form-control" type="file" name="file" required>
                </div>
                <button class="btn btn-primary" type="submit">Upload File</button>
            </form>
        </body>
        <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>                          
    </html>
    """, message=message)





@app.route('/delete/<row_id>', methods=['POST'])
@login_required
def delete_row(row_id):
    try:
        supabase.table("products").delete().eq("id", row_id).execute()
        return redirect(url_for('search_form', msg='deleted'))
    except Exception as e:
        return f"Error deleting row: {str(e)}", 500

    
@app.route('/delete-by-file', methods=['POST'])
@login_required
def delete_by_file():
    filename = request.form.get('source_file')
    if not filename:
        return "No file name provided", 400
    try:
        # Delete from all relevant tables
        supabase.table("products").delete().eq("source_file", filename).execute()
        supabase.table("packlist").delete().eq("source_file", filename).execute()
        supabase.table("bills").delete().eq("source_file", filename).execute()
        return redirect(url_for('search_form', msg='bulk_deleted', filename=filename))
    except Exception as e:
        return f"Error deleting rows from file: {str(e)}", 500



@app.route('/search-form', methods=['GET', 'POST'])
@login_required
def search_form():
    results = []
    query = ""
    message = request.args.get('msg')
    filename = request.args.get('filename')

    # Get unique Excel file names for dropdown
    file_list = supabase.table("products").select("source_file").execute().data
    unique_files = sorted(set(row['source_file'] for row in file_list if row.get('source_file')))

    if request.method == 'POST' and 'query' in request.form:
        query = request.form.get('query', '').strip()

        if query:
            code_match = supabase.table("products").select("*").ilike("code", f"%{query}%").execute().data
            desc_match = supabase.table("products").select("*").ilike("description", f"%{query}%").execute().data

            seen = set()
            for row in code_match + desc_match:
                row_id = row["id"]
                if row_id not in seen:
                    results.append(row)
                    seen.add(row_id)

    # Lookup buyers from 'bills' table
    bill_rows = supabase.table("bills").select("row_data").execute().data
    buyers_by_code = {}

    for bill in bill_rows:
        desc = str(bill['row_data'].get('Description', '')).strip()
        name = str(bill['row_data'].get('Name', '')).strip()
        price = bill['row_data'].get('Price', '')

        match = re.search(r'Code[:\s]*(\d+)', desc)
        if match:
            code = match.group(1).strip()
            buyers_by_code.setdefault(code, set()).add((name, price))

    # Convert sets to list of dicts for display
    for code in buyers_by_code:
        buyers_by_code[code] = [{'name': name, 'price': price} for name, price in buyers_by_code[code]]


    for r in results:
        r['buyers'] = buyers_by_code.get(r.get('code', ''), [])

    return render_template_string("""
    <!DOCTYPE html>
<html>
<head>
    <title>Search Products</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
</head>
<body class="container py-5">
    <nav class="navbar navbar-expand-lg navbar-dark bg-dark mb-4">
        <div class="container-fluid">
            <a class="navbar-brand" href="/search-form">🧾 CPSApp</a>
            <button class="navbar-toggler" type="button" data-bs-toggle="collapse" data-bs-target="#mainNavbar" aria-controls="mainNavbar" aria-expanded="false" aria-label="Toggle navigation">
            <span class="navbar-toggler-icon"></span>
            </button>
            
            <div class="collapse navbar-collapse" id="mainNavbar">
            <ul class="navbar-nav me-auto mb-2 mb-lg-0">
                <li class="nav-item"><a class="nav-link {% if request.path == '/upload-form' %}active{% endif %}" href="/upload-form">📤 Upload</a></li>
                <li class="nav-item"><a class="nav-link {% if request.path == '/search-form' %}active{% endif %}" href="/search-form">🔍 Search & Delete</a></li>
                <li class="nav-item"><a class="nav-link {% if request.path == '/view-packlist' %}active{% endif %}" href="/view-packlist">📦 Pack_List</a></li>
                <li class="nav-item"><a class="nav-link {% if request.path == '/invoice' %}active{% endif %}" href="/invoice">🧾 Invoice</a></li>
                <li class="nav-item"><a class="nav-link {% if request.path == '/generate-invoice' %}active{% endif %}" href="/generate-invoice">🧾 Generate Invoice</a></li>
            </ul>
            {% if session.get("user") %}
            <div class="d-flex align-items-center">
                <span class="navbar-text text-white me-3">👋 {{ session['user'] }}</span>
                <a href="/logout" class="btn btn-outline-light btn-sm">Logout</a>
            </div>
            {% endif %}
            </div>
        </div>
    </nav>

    <h2 class="mb-4">🔎 Search Products</h2>

    {% if message == 'deleted' %}
        <div class="alert alert-success">✅ Row successfully deleted.</div>
    {% elif message == 'bulk_deleted' %}
        <div class="alert alert-success">✅ All rows from file <strong>{{ filename }}</strong> were deleted.</div>
    {% endif %}

    <!-- 🔍 Search form -->
    <form method="post" class="mb-4">
        <div class="input-group">
            <input type="text" class="form-control" name="query" placeholder="Enter code or description" value="{{ query }}" required>
            <button class="btn btn-primary" type="submit">Search</button>
        </div>
    </form>

    <!-- 📂 Delete-by-file form -->
    <form method="post" action="/delete-by-file" class="mb-4">
        <div class="row g-2 align-items-center">
            <div class="col-auto">
                <select name="source_file" class="form-select" required>
                    <option value="">-- Select Excel file to delete --</option>
                    {% for f in unique_files %}
                        <option value="{{ f }}">{{ f }}</option>
                    {% endfor %}
                </select>
            </div>
            <div class="col-auto">
                <button class="btn btn-outline-danger" type="submit" onclick="return confirm('Delete all rows from this file?');">
                    Delete All From File
                </button>
            </div>
        </div>
    </form>

    {% if results %}
        <div class="table-responsive">
            <table class="table table-bordered align-middle">
                <thead class="table-light">
                    <tr>
                        <th>Code</th>
                        <th>Description</th>
                        <th>Price</th>
                        <th>Color</th>
                        <th>Measurements</th>
                        <th>Design Number</th>
                        <th>Source File</th>
                        <th>Buyers</th>
                        <th>Action</th>
                    </tr>
                </thead>
                <tbody>
                    {% for row in results %}
                        <tr>
                            <td>{{ row['code'] }}</td>
                            <td>{{ row['description'] }}</td>
                            <td>{{ row['price'] }}</td>
                            <td>{{ row['color'] }}</td>
                            <td>{{ row['measurements'] }}</td>
                            <td>{{ row['design_number'] }}</td>
                            <td>{{ row['source_file'] }}</td>
                            <td>
                                {% if row['buyers'] %}
                                    <ul class="mb-0">
                                        {% for buyer in row['buyers'] %}
                                            <li>{{ buyer.name }}</li>
                                        {% endfor %}
                                    </ul>
                                {% else %}
                                    <span class="text-muted">No buyers</span>
                                {% endif %}
                            </td>
                            <td>
                                <form method="post" action="/delete/{{ row['id'] }}" onsubmit="return confirm('Delete this row?');">
                                    <button class="btn btn-sm btn-danger">Delete</button>
                                </form>
                            </td>
                        </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
    {% elif query %}
        <div class="alert alert-warning">No results found for "{{ query }}".</div>
    {% endif %}
</body>
</html>
<script>
    document.addEventListener('DOMContentLoaded', function() {
        const searchInput = document.querySelector('input[name="query"]');
        searchInput.blur();  // Remove focus after page load
    });
</script>
                                  

    """, results=results, query=query, message=message, filename=filename, unique_files=unique_files)




def process_product_tab(sheet, product_name):
    # Example: existing logic to extract other product info...

    measurements = extract_measurements(sheet)

    product_data = {
        'name': product_name,
        # 'other_field': value,
        'measurements': measurements,
    }

    return product_data

@app.route('/get-products')
def get_products():
    products = supabase.table('products').select("*").execute().data

    filtered = []
    for p in products:
        # Remove ID and Upload Date
        p.pop('id', None)
        p.pop('upload_date', None)
        filtered.append(p)

    return jsonify(filtered)


@app.route('/view-packlist', methods=['GET', 'POST'])
@login_required
def view_packlist():
    selected_file = None
    packlist_df = None
    error = None

    # Fetch file list from Supabase
    file_rows = supabase.table("packlist").select("source_file").execute().data
    unique_files = sorted({r["source_file"] for r in file_rows if r.get("source_file")}, reverse=True)

    if request.method == 'POST':
        selected_file = request.form.get('selected_file')
        try:
            rows = supabase.table("packlist").select("row_data").eq("source_file", selected_file).order("row_index").execute().data
            if rows:
                packlist_df = pd.DataFrame([row['row_data'] for row in rows])
            else:
                error = f"❌ No Pack_List data found for file: {selected_file}"
        except Exception as e:
            error = f"❌ Error loading Pack_List from Supabase: {str(e)}"

    return render_template_string("""
    <html>
        <head>
            <title>📦 View Pack_List</title>
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
        </head>
        <body class="container py-5">
            <nav class="navbar navbar-expand-lg navbar-dark bg-dark mb-4">
                <div class="container-fluid">
                    <a class="navbar-brand" href="/search-form">🧾 CPSApp</a>
                    <button class="navbar-toggler" type="button" data-bs-toggle="collapse" data-bs-target="#mainNavbar" aria-controls="mainNavbar" aria-expanded="false" aria-label="Toggle navigation">
                    <span class="navbar-toggler-icon"></span>
                    </button>
                    
                    <div class="collapse navbar-collapse" id="mainNavbar">
                    <ul class="navbar-nav me-auto mb-2 mb-lg-0">
                        <li class="nav-item"><a class="nav-link {% if request.path == '/upload-form' %}active{% endif %}" href="/upload-form">📤 Upload</a></li>
                        <li class="nav-item"><a class="nav-link {% if request.path == '/search-form' %}active{% endif %}" href="/search-form">🔍 Search & Delete</a></li>
                        <li class="nav-item"><a class="nav-link {% if request.path == '/view-packlist' %}active{% endif %}" href="/view-packlist">📦 Pack_List</a></li>
                        <li class="nav-item"><a class="nav-link {% if request.path == '/invoice' %}active{% endif %}" href="/invoice">🧾 Invoice</a></li>
                        <li class="nav-item"><a class="nav-link {% if request.path == '/generate-invoice' %}active{% endif %}" href="/generate-invoice">🧾 Generate Invoice</a></li>
                    </ul>
                    {% if session.get("user") %}
                    <div class="d-flex align-items-center">
                        <span class="navbar-text text-white me-3">👋 {{ session['user'] }}</span>
                        <a href="/logout" class="btn btn-outline-light btn-sm">Logout</a>
                    </div>
                    {% endif %}
                    </div>
                </div>
            </nav>


            <h2 class="mb-4">📦 View Pack_List Sheet</h2>

            <form method="post" class="mb-4">
                <div class="input-group">
                    <select name="selected_file" class="form-select" required>
                        <option value="">-- Select uploaded file --</option>
                        {% for file in unique_files %}
                            <option value="{{ file }}" {% if file == selected_file %}selected{% endif %}>{{ file }}</option>
                        {% endfor %}
                    </select>
                    <button class="btn btn-primary" type="submit">View Pack_List</button>
                </div>
            </form>

            {% if error %}
                <div class="alert alert-danger">{{ error }}</div>
            {% endif %}

            {% if packlist_df is not none %}
                <div class="table-responsive">
                    <table class="table table-bordered table-striped">
                        <thead class="table-light">
                            <tr>
                                {% for col in packlist_df.columns %}
                                    <th>{{ col }}</th>
                                {% endfor %}
                            </tr>
                        </thead>
                        <tbody>
                            {% for _, row in packlist_df.iterrows() %}
                                <tr>
                                    {% for cell in row %}
                                        <td>{{ cell }}</td>
                                    {% endfor %}
                                </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </div>
            {% endif %}
        </body>
    </html>
    """, unique_files=unique_files, selected_file=selected_file, packlist_df=packlist_df, error=error)

@app.route('/invoice', methods=['GET', 'POST'])
@login_required
def invoice():
    invoice_number = f"INV-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
    invoice_date = datetime.utcnow().strftime('%Y-%m-%d')
    live_session_number = ""
    products = []
    courier_fee = 0
    total_amount = 0
    delivery_method = ""
    show_invoice = False

    if request.method == 'POST':
        invoice_number = request.form.get('invoice_number')
        invoice_date = request.form.get('invoice_date')
        live_session_number = request.form.get('live_session_number')
        products_raw = request.form.get('products', '')
        delivery_method = request.form.get('delivery_method')

        # Parse Products
        for line in products_raw.strip().split('\n'):
            if line.strip():
                try:
                    desc, code, price, qty = [p.strip() for p in line.split('|')]
                    price = float(price)
                    qty = int(qty)
                    subtotal = price * qty
                    products.append({
                        'desc': desc,
                        'code': code,
                        'price': price,
                        'qty': qty,
                        'subtotal': subtotal
                    })
                    total_amount += subtotal
                except Exception as e:
                    print(f"[ERROR] Failed to parse line: {line}, Error: {e}", flush=True)

        if delivery_method == 'Courier Service':
            courier_fee = 4
            total_amount += courier_fee

        show_invoice = True

    return render_template_string("""
    <html>
    <head>
        <title>Create Invoice</title>
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    </head>
    <body class="container py-5">
       <nav class="navbar navbar-expand-lg navbar-dark bg-dark mb-4">
            <div class="container-fluid">
                <a class="navbar-brand" href="/search-form">🧾 CPSApp</a>
                <button class="navbar-toggler" type="button" data-bs-toggle="collapse" data-bs-target="#mainNavbar" aria-controls="mainNavbar" aria-expanded="false" aria-label="Toggle navigation">
                <span class="navbar-toggler-icon"></span>
                </button>
                
                <div class="collapse navbar-collapse" id="mainNavbar">
                <ul class="navbar-nav me-auto mb-2 mb-lg-0">
                    <li class="nav-item"><a class="nav-link {% if request.path == '/upload-form' %}active{% endif %}" href="/upload-form">📤 Upload</a></li>
                    <li class="nav-item"><a class="nav-link {% if request.path == '/search-form' %}active{% endif %}" href="/search-form">🔍 Search & Delete</a></li>
                    <li class="nav-item"><a class="nav-link {% if request.path == '/view-packlist' %}active{% endif %}" href="/view-packlist">📦 Pack_List</a></li>
                    <li class="nav-item"><a class="nav-link {% if request.path == '/invoice' %}active{% endif %}" href="/invoice">🧾 Invoice</a></li>
                    <li class="nav-item"><a class="nav-link {% if request.path == '/generate-invoice' %}active{% endif %}" href="/generate-invoice">🧾 Generate Invoice</a></li>
                </ul>
                {% if session.get("user") %}
                <div class="d-flex align-items-center">
                    <span class="navbar-text text-white me-3">👋 {{ session['user'] }}</span>
                    <a href="/logout" class="btn btn-outline-light btn-sm">Logout</a>
                </div>
                {% endif %}
                </div>
            </div>
    </nav>

        </nav>

        <h2 class="mb-4">🧾 Create Invoice</h2>
        <form method="post">
            <div class="mb-3">
                <label>Invoice Number</label>
                <input type="text" name="invoice_number" class="form-control" value="{{ invoice_number }}" readonly>
            </div>
            <div class="mb-3">
                <label>Invoice Date</label>
                <input type="text" name="invoice_date" class="form-control" value="{{ invoice_date }}" readonly>
            </div>
            <div class="mb-3">
                <label>Live Session Number</label>
                <input type="text" name="live_session_number" class="form-control" value="{{ live_session_number }}">
            </div>
            <div class="mb-3">
                <label>Products</label>
                <textarea name="products" class="form-control" rows="5" placeholder="Format: Description | Code | Price | Qty"></textarea>
                <small class="text-muted">One line per product, separated by |</small>
            </div>
            <div class="mb-3">
                <label>Delivery Method</label>
                <div class="form-check">
                    <input class="form-check-input" type="radio" name="delivery_method" value="Courier Service" id="courier" required>
                    <label class="form-check-label" for="courier">Courier Service (+$4)</label>
                </div>
                <div class="form-check">
                    <input class="form-check-input" type="radio" name="delivery_method" value="Self Collection" id="self">
                    <label class="form-check-label" for="self">Self Collection</label>
                </div>
                <div class="form-check">
                    <input class="form-check-input" type="radio" name="delivery_method" value="Accumulation" id="accumulation">
                    <label class="form-check-label" for="accumulation">Accumulation</label>
                </div>
            </div>
            <button type="submit" class="btn btn-primary">Generate Invoice</button>
        </form>

        {% if show_invoice %}
            <h3 class="mt-5">Invoice</h3>
            <p><strong>Invoice Number:</strong> {{ invoice_number }}</p>
            <p><strong>Invoice Date:</strong> {{ invoice_date }}</p>
            <p><strong>Live Session:</strong> {{ live_session_number }}</p>
            <table class="table table-bordered">
                <thead><tr><th>Description</th><th>Code</th><th>Price</th><th>Qty</th><th>Subtotal</th></tr></thead>
                <tbody>
                {% for p in products %}
                    <tr>
                        <td>{{ p.desc }}</td>
                        <td>{{ p.code }}</td>
                        <td>${{ "%.2f"|format(p.price) }}</td>
                        <td>{{ p.qty }}</td>
                        <td>${{ "%.2f"|format(p.subtotal) }}</td>
                    </tr>
                {% endfor %}
                    {% if courier_fee %}
                    <tr><td colspan="4"><strong>Courier Fee</strong></td><td>${{ courier_fee }}</td></tr>
                    {% endif %}
                    <tr><td colspan="4"><strong>Total</strong></td><td><strong>${{ total_amount }}</strong></td></tr>
                </tbody>
            </table>

            <div class="alert alert-info">
                <strong>Payment Instructions:</strong><br>
                1. Bank transfer to OCBC current account 588056739001<br>
                2. PAYNOW to UEN number: 201013470W<br>
                Cupid Apparel Pte Ltd<br><br>
                <strong>** Kindly indicate your FB name in the payment description, and screenshot your payment.</strong>
            </div>
        {% endif %}
    </body>
    </html>
    """, invoice_number=invoice_number, invoice_date=invoice_date, live_session_number=live_session_number,
       products=products, courier_fee=courier_fee, total_amount=total_amount,
       show_invoice=show_invoice)


from flask import request, render_template_string, session
from datetime import datetime
import json

@app.route('/generate-invoice', methods=['GET', 'POST'])
@login_required

from flask import request, render_template_string, session
from datetime import datetime
import json

@app.route('/generate-invoice', methods=['GET', 'POST'])
@login_required

def generate_invoice():
    # 0) Grab posted values
    selected_file     = request.form.get("selected_file", "")
    selected_customer = request.form.get("selected_customer", "")
    invoice_data      = None
    adhoc_items       = []

    # 1) Load & group Bill-tab rows
    file_rows = supabase.table("bills")\
                       .select("source_file, row_data")\
                       .execute().data
    file_to_rows = {}
    for r in file_rows:
        f = r.get("source_file")
        d = r.get("row_data") or {}
        file_to_rows.setdefault(f, []).append(d)
    unique_files = sorted(file_to_rows.keys(), reverse=True)

    # 2) Build customer list
    all_rows = file_to_rows.get(selected_file, [])
    customer_to_rows = {}
    for d in all_rows:
        name = (d.get("Name") or "").strip()
        if name:
            customer_to_rows.setdefault(name, []).append(d)
    customer_list = sorted(customer_to_rows.keys())

    # 3) Pull Master-tab product catalogue, group by design_number
    raw_prods = supabase.table("products")\
                        .select("id,design_number,code,description,price,color")\
                        .order("design_number")\
                        .execute().data
    products = [{
        "id":            str(p["id"]),
        "design_number": str(p["design_number"]),
        "code":          p["code"],
        "description":   p["description"],
        "price":         float(p["price"] or 0),
        "color":         p["color"]
    } for p in raw_prods]

    # build design_map and flat designs list
    design_map = {}
    for sku in products:
        design_map.setdefault(sku["design_number"], []).append(sku)

    designs = [
        {"design_number": dn, "description": design_map[dn][0]["description"]}
        for dn in sorted(design_map.keys(), key=lambda x: int(x))
    ]

    design_map_json = json.dumps(design_map)  # for client-side JS

    # 4) If file & customer chosen → build invoice_data
    if selected_file and selected_customer:
        # a) Aggregate Bill-tab purchases
        items_map = {}
        subtotal = total_qty = 0
        for d in customer_to_rows[selected_customer]:
            desc  = (d.get("Description") or "").strip()
            price = float(d.get("Price") or 0)
            key   = (desc, price)
            items_map.setdefault(key, {"Description": desc, "Price": price, "Qty": 0})
            items_map[key]["Qty"] += 1
            subtotal += price
            total_qty += 1
        items = list(items_map.values())

        # b) Read & remember all ad-hoc rows
        designs_post = request.form.getlist("item_design")
        sku_ids      = request.form.getlist("item_sku")
        descs        = request.form.getlist("item_desc")
        qtys         = request.form.getlist("item_qty")
        prices       = request.form.getlist("item_price")

        for design, sku_id, desc, q, p in zip(
            designs_post, sku_ids, descs, qtys, prices
        ):
            # skip fully blank rows
            if not (sku_id or (design == "other" and desc.strip())):
                continue
            qty = int(q or 0)
            if sku_id and sku_id != "other":
                sku = next(x for x in products if x["id"] == sku_id)
                price = sku["price"]
                label = (f"{sku['design_number']} | {sku['code']} | "
                         f"{sku['description']} | {sku['color']}")
                desc  = sku["description"]
            else:
                # free-text
                price = float(p or 0)
                label = desc

            adhoc_items.append({
                "design": design,
                "sku":    sku_id,
                "desc":   desc,
                "qty":    qty,
                "price":  price
            })
            items.append({"Description": label, "Price": price, "Qty": qty})
            subtotal += price * qty
            total_qty += qty

        # c) Courier logic
        method = request.form.get("courier_method", "")
        outlet = request.form.get("outlet_option", "")
        fee    = 4 if method == "Courier Service" else 0
        courier_label = (
            f"Self Collection - {outlet}"
            if method == "Self Collection" and outlet else
            method
        )
        total = subtotal + fee

        # d) Build invoice_text & invoice_data
        lines = [
            f"Hi {selected_customer},",
            "Thank you for your support.", "",
            f"Live Session: {selected_file}",
            f"Date: {datetime.utcnow():%Y-%m-%d}", "",
            "Items:"
        ]
        for it in items:
            lines.append(
                f"- {it['Description']} | ${it['Price']} x {it['Qty']} = "
                f"${it['Price']*it['Qty']:.2f}"
            )
        lines += [
            f"Total Quantity: {total_qty}",
            f"Subtotal: ${subtotal:.2f}",
            f"Courier Fee: ${fee:.2f}",
            f"Total: ${total:.2f}",
            f"Courier Method: {courier_label}", "",
            "Please make payment via:",
            "1. Bank transfer to OCBC current account 588056739001",
            "2. PAYNOW to UEN number: 201013470W",
            "Cupid Apparel Pte Ltd",
            "** Kindly indicate your FB name in the payment description, and screenshot your payment"
        ]
        invoice_data = {
            "customer":             selected_customer,
            "file":                 selected_file,
            "invoice_date":         datetime.utcnow().strftime("%Y-%m-%d"),
            "items":                items,
            "total_quantity":       total_qty,
            "subtotal":             subtotal,
            "courier_fee":          fee,
            "total":                total,
            "courier":              courier_label,
            "payment_instructions": "\n".join(lines[-6:]),
            "invoice_text":         "\n".join(lines)
        }

    # 5) Render form + preview
    return render_template_string("""
<!doctype html>
<html lang="en"><head>
  <meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
  <script src="https://cdn.jsdelivr.net/npm/html2canvas@1.4.1/dist/html2canvas.min.js"></script>
  <style>.remove-item{font-size:1.2rem}</style>
</head><body class="container py-5">

  <h2 class="mb-4">🧾 Generate Invoice</h2>
  <form method="post" id="invoice-form">
    <!-- 1) File -->
    <div class="mb-3">
      <label>Select File</label>
      <select name="selected_file" class="form-select" onchange="this.form.submit()">
        <option value="">-- choose session --</option>
        {% for f in unique_files %}
          <option value="{{f}}" {% if f==selected_file %}selected{% endif %}>{{f}}</option>
        {% endfor %}
      </select>
    </div>

    {% if selected_file %}
    <!-- 2) Customer -->
    <div class="mb-3">
      <label>Select Customer</label>
      <select name="selected_customer" class="form-select" onchange="this.form.submit()">
        <option value="">-- choose customer --</option>
        {% for c in customer_list %}
          <option value="{{c}}" {% if c==selected_customer %}selected{% endif %}>{{c}}</option>
        {% endfor %}
      </select>
    </div>
    {% endif %}

    {% if selected_customer %}
    <!-- 3) Courier -->
    <div class="mb-3">
      <label>Courier Method</label><br>
      <label>
        <input type="radio" name="courier_method" value="Courier Service" required
               onchange="this.form.submit()"
               {% if invoice_data and invoice_data.courier=='Courier Service' %}checked{% endif %}>
        Courier (+$4)
      </label><br>
      <label>
        <input type="radio" name="courier_method" value="Self Collection"
               onchange="this.form.submit()"
               {% if invoice_data and invoice_data.courier.startswith('Self Collection') %}checked{% endif %}>
        Self Collection
      </label>
      <select name="outlet_option" class="form-select mt-2" onchange="this.form.submit()">
        <option value="">-- Outlet --</option>
        <option>Westmall</option><option>Jurong Point 2</option><option>Northpoint City</option>
      </select><br>
      <label>
        <input type="radio" name="courier_method" value="Accumulation"
               onchange="this.form.submit()"
               {% if invoice_data and invoice_data.courier=='Accumulation' %}checked{% endif %}>
        Accumulation
      </label>
    </div>

    <!-- 4) Ad-hoc Items -->
    <h5 class="mt-4">Add Ad-hoc Items</h5>
    <table class="table" id="items-table">
      <thead><tr>
        <th>Product</th><th>Color</th><th>Description</th><th>Qty</th><th>Unit Price</th><th></th>
      </tr></thead>
      <tbody>
        {% set rows = adhoc_items or [{}] %}
        {% for row in rows %}
        <tr class="item-row">
          <td>
            <select name="item_design" class="form-select design-select">
              <option value="">-- select product --</option>
              {% for d in designs %}
                <option value="{{d.design_number}}"
                  {% if row.design==d.design_number %}selected{% endif %}>
                  {{d.design_number}} – {{d.description}}
                </option>
              {% endfor %}
              <option value="other" {% if row.sku=='other' %}selected{% endif %}>
                Free Text…
              </option>
            </select>
          </td>
          <td>
            <select name="item_sku" class="form-select sku-select" disabled>
              <option value="">-- select color --</option>
            </select>
          </td>
          <td>
            <input type="text" name="item_desc" class="form-control desc-input"
                   placeholder="Description"
                   {% if row.design!='other' %}readonly{% endif %}
                   value="{{row.desc or ''}}">
          </td>
          <td><input type="number" name="item_qty" class="form-control qty-input" min="1"
                     value="{{row.qty or 1}}"></td>
          <td><input type="number" name="item_price" class="form-control price-input" step="0.01"
                     readonly value="{{row.price or ''}}"></td>
          <td><button type="button" class="btn btn-outline-danger btn-sm remove-item">×</button></td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
    <button type="button" id="add-item" class="btn btn-sm btn-outline-primary mb-3">
      Add Item
    </button>
    {% endif %}
  </form>

  <script>
    const designMap = {{ design_map_json|safe }};
    const invoiceForm = document.getElementById('invoice-form');

    function attachHandlers(row) {
      const designSel = row.querySelector('.design-select');
      const skuSel    = row.querySelector('.sku-select');
      const descIn    = row.querySelector('.desc-input');
      const qtyIn     = row.querySelector('.qty-input');
      const priceIn   = row.querySelector('.price-input');

      designSel.onchange = () => {
        const d = designSel.value;
        skuSel.innerHTML = '<option value="">-- select color --</option>';
        skuSel.disabled  = true;
        descIn.value     = '';
        descIn.readOnly  = true;
        priceIn.value    = '';
        qtyIn.value      = 1;

        if (designMap[d]) {
          skuSel.disabled = false;
          designMap[d].forEach(sku => {
            const opt = document.createElement('option');
            opt.value        = sku.id;
            opt.dataset.price = sku.price;
            opt.textContent   = sku.color;
            skuSel.appendChild(opt);
          });
        }
        if (d === 'other') descIn.readOnly = false;
        invoiceForm.submit();
      };

      skuSel.onchange = () => {
        const o = skuSel.selectedOptions[0];
        const sku = designMap[designSel.value].find(s=>s.id===o.value);
        priceIn.value   = sku.price;
        descIn.value    = sku.description;
        descIn.readOnly = true;
        invoiceForm.submit();
      };

      descIn.onchange = qtyIn.onchange = () => invoiceForm.submit();
    }

    document.getElementById('add-item').onclick = () => {
      const tbody = document.querySelector('#items-table tbody');
      const proto = tbody.querySelector('.item-row');
      const nr    = proto.cloneNode(true);
      nr.querySelectorAll('select, input').forEach(el=>el.value='');
      nr.querySelector('.qty-input').value     = 1;
      nr.querySelector('.desc-input').readOnly = true;
      nr.querySelector('.sku-select').disabled = true;
      nr.querySelector('.sku-select').innerHTML= '<option>-- select color --</option>';
      tbody.appendChild(nr);
      attachHandlers(nr);
    };

    document.querySelector('#items-table').addEventListener('click', e=>{
      if (e.target.matches('.remove-item')) {
        const rows = document.querySelectorAll('.item-row');
        if (rows.length>1) e.target.closest('tr').remove();
        invoiceForm.submit();
      }
    });

    // bind on load
    document.querySelectorAll('.item-row').forEach(attachHandlers);
  </script>

  {% if invoice_data %}
  <hr><h4>Invoice Preview</h4>
  <p><strong>Hi:</strong> {{invoice_data.customer}}</p>
  <p><strong>Session:</strong> {{invoice_data.file}}</p>
  <p><strong>Date:</strong> {{invoice_data.invoice_date}}</p>
  <table class="table table-bordered">
    <thead><tr><th>Description</th><th>Price</th><th>Qty</th><th>Subtotal</th></tr></thead>
    <tbody>
      {% for it in invoice_data['items'] %}
      <tr>
        <td>{{it.Description}}</td>
        <td>${{ '%.2f'|format(it.Price) }}</td>
        <td>{{it.Qty}}</td>
        <td>${{ '%.2f'|format(it.Price * it.Qty) }}</td>
      </tr>
      {% endfor %}
      <tr><td colspan="3"><strong>Total Quantity</strong></td><td>{{invoice_data.total_quantity}}</td></tr>
      <tr><td colspan="3"><strong>Subtotal</strong></td><td>${{ '%.2f'|format(invoice_data.subtotal) }}</td></tr>
      <tr><td colspan="3"><strong>Courier Fee</strong></td><td>${{ '%.2f'|format(invoice_data.courier_fee) }}</td></tr>
      <tr><td colspan="3"><strong>Total</strong></td><td><strong>${{ '%.2f'|format(invoice_data.total) }}</strong></td></tr>
    </tbody>
  </table>
  <p><strong>Courier Method:</strong> {{invoice_data.courier}}</p>
  <div class="alert alert-info">
    {{invoice_data.payment_instructions.replace('\n','<br>')|safe}}
  </div>
  {% endif %}

</body></html>
    """,
    unique_files      = unique_files,
    selected_file     = selected_file,
    selected_customer = selected_customer,
    customer_list     = customer_list,
    designs           = designs,
    design_map        = design_map,
    design_map_json   = design_map_json,
    adhoc_items       = adhoc_items,
    invoice_data      = invoice_data
)





if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5002))
    app.run(host='0.0.0.0', port=port)
